from openpyxl import load_workbook
import time
from geopy.geocoders import Nominatim
from geopy.distance import geodesic
import folium
import os

fn = "C:/Users/holog/OneDrive/Рабочий стол/GPS/Analyzer/gds.xlsx"
wb = load_workbook(fn)
ws = wb["data"]

gps_data = open("C:/Users/holog/OneDrive/Рабочий стол/GPS/Analyzer/putty.txt", "r").read()
parsed_string = gps_data.split("\n")
true_parsed = []

pars_index = 0
data_gps = ""
data_temp = ""
data_hum = ""

def clear_console():
    os.system('cls' if os.name == 'nt' else 'clear')

print("Data processing...")


for veo in parsed_string:

	if pars_index == 0:
		if veo.startswith("GPRMC") and veo.count(",") == 12 and ",,," in veo:
			data_gps = veo
			pars_index = 1

	elif pars_index == 1:
		if veo.startswith("TEMP") and len(veo) > 6:
			data_temp = veo
			pars_index = 2
		else:
			pars_index = 0

	elif pars_index == 2:
		if veo.startswith("HUM") and len(veo) > 5:
			data_hum = veo
			pars_index = 3
			true_parsed.append(data_gps)
			true_parsed.append(data_temp)
			true_parsed.append(data_hum)
			pars_index = 0
		else:
			pars_index = 0

alpars = len(true_parsed) 
only_20 = (len(true_parsed) // 60) 
if only_20 == 0:
	only_20 = 1

ws["P3"] = round(((len(true_parsed) / len(parsed_string)) * 100), 2)
ws["P7"] = len(true_parsed) // 3

parsed_20 = []
tripple = len(true_parsed) // 3 
if tripple > 20:
	tripple = 20
coun_now = 0
sir = 0

for veo in range(tripple):
	sir = 0
	while sir < 3:
		parsed_20.append(true_parsed[(3 * coun_now * only_20 + sir)])
		sir += 1
	coun_now +=1

column = 3
act_lattitude = 0
act_longitude = 0
time_now = ""
geolocator = Nominatim(user_agent = "Hologram")
procent = 5

xsl_index = 0
for veo in parsed_20:

	if xsl_index == 0:
		ws["C" + str(column)] = column - 2

		red_veo = (veo[0:(veo.index(",,,"))]).replace(",,", ",")
		parameters = red_veo.split(",")
		par_index = 0

		for i in parameters:
			if par_index == 0:
				ws["D" + str(column)] = i
			elif par_index == 1:
				ws["E" + str(column)] = str(i[0:2] + ":" + i[2:4] + ":" + i[4:6])
				time_now = str(i[0:2] + "." + i[2:4] + "." + i[4:6])
			elif par_index == 3:
				ws["F" + str(column)] = float(str(i[0:2] + "." + i[2:4] + i[5::]))
				act_lattitude = float(str(i[0:2] + "." + i[2:4] + i[5::]))
			elif par_index == 4:
				ws["G" + str(column)] = i
			elif par_index == 5:
				ws["H" + str(column)] = float(str(i[1:3] + "." + i[3:5] + i[6::]))
				act_longitude = float(str(i[1:3] + "." + i[3:5] + i[6::]))

				geo = ((geolocator.reverse(str(act_lattitude) + ", " + str(act_longitude))).address).split(", ")
				ws["S" + str(column)] = geo[5]
				ws["T" + str(column)] = geo[1]

				ws["V" + str(column)] = geodesic((act_lattitude, act_longitude), (49.2257511, 16.5756361)).km
				ws["W" + str(column)] = geodesic((act_lattitude, act_longitude), (50.0596288, 14.446459273258009)).km
				ws["X" + str(column)] = geodesic((act_lattitude, act_longitude), (51.5074456, -0.1277653)).km

				m = folium.Map(location=[act_lattitude, act_longitude], zoom_start=12)
				folium.Marker(location=[act_lattitude, act_longitude]).add_to(m)
				m.save('C:/Users/holog/OneDrive/Рабочий стол/GPS/Analyzer/maps/' + time_now + '.html')

			elif par_index == 6:
				ws["I" + str(column)] = i
			elif par_index == 7:
				ws["J" + str(column)] = float(i)
			elif par_index == 8:
				ws["K" + str(column)] = str(i[0:2] + "." + i[2:4] + "." + i[4:6])
			par_index += 1
		xsl_index = 1

	elif xsl_index == 1:
		ws["M" + str(column)] = float(str(veo[6::]))
		xsl_index = 2

	elif xsl_index == 2:
		ws["N" + str(column)] = float(str(veo[5::]))
		xsl_index = 0
		column += 1
		print(str(procent)+"%")
		procent += 5

wb.save(fn)
wb.close()
clear_console()
print("\033[32mData processing completed successfully.\033[0m")
time.sleep(3)